#!/usr/bin/env python3
import boto3
import json
import argparse
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class AWSResourceExplorer:
    def __init__(self, region='us-east-1'):
        self.region = region
        self.session = boto3.Session()
    
    def get_resources(self):
        try:
            client = self.session.client('resource-explorer-2', region_name=self.region)
            
            # Check if Resource Explorer is available
            try:
                client.list_indexes()
            except client.exceptions.ResourceNotFoundException:
                print(f"✗ Resource Explorer not enabled in {self.region}")
                return []
            
            resources = []
            
            # Search for all resources
            response = client.search(
                QueryString='*', 
                MaxResults=1000
            )
            
            for resource in response.get('Resources', []):
                arn = resource.get('Arn', '')
                properties = resource.get('Properties', {})
                
                # Extract application from tags or properties
                application = ''
                if properties:
                    # Check for Application tag in various formats
                    if 'Tags' in properties:
                        tags = properties['Tags']
                        if isinstance(tags, dict):
                            application = (tags.get('Application') or 
                                         tags.get('application') or 
                                         tags.get('App') or 
                                         tags.get('app') or '')
                        elif isinstance(tags, list):
                            for tag in tags:
                                if isinstance(tag, dict) and tag.get('Key', '').lower() in ['application', 'app']:
                                    application = tag.get('Value', '')
                                    break
                
                resources.append({
                    'Service': resource.get('Service', ''),
                    'ResourceType': resource.get('ResourceType', ''),
                    'Identifier': arn.split('/')[-1] if '/' in arn else arn.split(':')[-1] if ':' in arn else arn,
                    'ARN': arn,
                    'Application': application,
                    'AWSAccount': resource.get('OwningAccountId', ''),
                    'Region': resource.get('Region', ''),
                    'LastReportedAt': resource.get('LastReportedAt', '').isoformat() if resource.get('LastReportedAt') else ''
                })
            
            print(f"✓ Found {len(resources)} resources in {self.region}")
            return resources
            
        except Exception as e:
            print(f"✗ Error in {self.region}: {str(e)}")
            return []

def get_all_regions():
    ec2 = boto3.client('ec2', region_name='us-east-1')
    return [r['RegionName'] for r in ec2.describe_regions()['Regions']]

def multi_region_inventory(regions=None):
    if not regions:
        regions = get_all_regions()
    
    all_resources = []
    
    def process_region(region):
        print(f"\nProcessing {region}...")
        return AWSResourceExplorer(region).get_resources()
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        results = executor.map(process_region, regions)
        for region_resources in results:
            all_resources.extend(region_resources)
    
    return all_resources

def save_json(resources, filename):
    inventory = {
        'timestamp': datetime.now().isoformat(),
        'total_resources': len(resources),
        'resources': resources
    }
    
    with open(filename, 'w') as f:
        json.dump(inventory, f, indent=2, default=str)
    print(f"\nJSON saved to {filename}")

def save_excel(resources, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "AWS Resources"
    
    # Headers
    headers = ['Service', 'ResourceType', 'Identifier', 'ARN', 'Application', 'AWSAccount', 'Region', 'LastReportedAt']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Data
    for row, resource in enumerate(resources, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=resource.get(header, ''))
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(filename)
    print(f"Excel saved to {filename}")

def upload_to_s3(filename, bucket_name, key_prefix="AWS_Inventory/"):
    """Upload file to S3 bucket with timestamp folder"""
    s3 = boto3.client('s3')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    key = f"{key_prefix}{timestamp}/{filename}"
    
    try:
        s3.upload_file(filename, bucket_name, key)
        print(f"✓ Uploaded {filename} to s3://{bucket_name}/{key}")
        return True
    except Exception as e:
        print(f"✗ S3 upload failed: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(description='AWS Resource Explorer Inventory')
    parser.add_argument('--region', '-r', help='Single region')
    parser.add_argument('--regions', nargs='+', help='Multiple regions')
    parser.add_argument('--all-regions', '-a', action='store_true', help='All regions')
    parser.add_argument('--output', '-o', default='aws_resources', help='Output filename (without extension)')
    parser.add_argument('--s3-bucket', help='S3 bucket name to upload files')
    
    args = parser.parse_args()
    
    if args.all_regions:
        resources = multi_region_inventory()
    elif args.regions:
        resources = multi_region_inventory(args.regions)
    else:
        region = args.region or 'us-east-1'
        resources = AWSResourceExplorer(region).get_resources()
    
    save_json(resources, f"{args.output}.json")
    save_excel(resources, f"{args.output}.xlsx")
    
    # Upload to S3 if bucket specified
    if args.s3_bucket:
        upload_to_s3(f"{args.output}.xlsx", args.s3_bucket)
        upload_to_s3(f"{args.output}.json", args.s3_bucket)
    
    print(f"\nTotal resources found: {len(resources)}")

if __name__ == "__main__":
    main()
